from __future__ import annotations

from typing import Any, Dict, List, Mapping, Optional

from openpyxl.workbook.workbook import Workbook

try:
    from spreadsheet_io.spreadsheet_utils import SHEET_NAME_TO_SPEC, normalize_spreadsheet_string_cell
except ModuleNotFoundError:
    from spreadsheet_utils import SHEET_NAME_TO_SPEC, normalize_spreadsheet_string_cell  # type: ignore[no-redef]

ENTITY_NOTE_SHEETS = (
    ("Sections", "sections", "id"),
    ("Instructors", "instructors", "id"),
    ("Rooms", "rooms", "id"),
    ("Timeslots", "timeslots", "id"),
    ("MeetingPatterns", "meeting-patterns", "id"),
)

def format_prev_notes_display(notes: List[Dict[str, Any]]) -> str:
    if not notes:
        return ""

    def _ts(note: Dict[str, Any]) -> str:
        return str(note.get("createdAt") or "")

    ordered = sorted(notes, key=_ts)
    blocks: List[str] = []

    for note in ordered:
        complete_tag = "[complete]" if note.get("completed") else ""
        blocks.append(
            f"[note_id={note.get('id', '')}][{note.get('createdAt', '')}]"
            f"[{note.get('author', '')}]{complete_tag}"
        )
        blocks.append(str(note.get("note") or "").strip())
        replies = sorted(note.get("replies") or [], key=lambda r: str(r.get("createdAt") or ""))
        for reply in replies:
            body = str(reply.get("note") or "").strip().replace("\n", "\n  ")
            blocks.append(
                f"  [reply_id={reply.get('id', '')}][{reply.get('createdAt', '')}]"
                f"[{reply.get('author', '')}]"
            )
            blocks.append(f"  {body}")
        blocks.append("")

    return "\n".join(blocks).strip()


def _header_column_index(ws, header_name: str) -> Optional[int]:
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if value is not None and str(value).strip().lower() == header_name.lower():
            return col_idx
    return None


def _set_column_values_for_keys(
    ws,
    header_name: str,
    values_by_row_key: Mapping[str, str],
    id_column: str,
) -> None:
    """Set column cells only for row keys present in values_by_row_key (do not blank other rows)."""
    if not values_by_row_key:
        return
    target_col = _header_column_index(ws, header_name)
    id_col = _header_column_index(ws, id_column)
    if not target_col or not id_col:
        return
    for row_idx in range(2, ws.max_row + 1):
        row_key = normalize_spreadsheet_string_cell(ws.cell(row=row_idx, column=id_col).value)
        if not row_key or row_key not in values_by_row_key:
            continue
        text = values_by_row_key[row_key]
        ws.cell(row=row_idx, column=target_col).value = text or None


def _clear_column(ws, header_name: str, id_column: str) -> None:
    target_col = _header_column_index(ws, header_name)
    id_col = _header_column_index(ws, id_column)
    if not target_col or not id_col:
        return
    for row_idx in range(2, ws.max_row + 1):
        row_key = normalize_spreadsheet_string_cell(ws.cell(row=row_idx, column=id_col).value)
        if not row_key:
            continue
        ws.cell(row=row_idx, column=target_col).value = None


def _notes_sheet_rows(note_entries: List[Dict[str, Any]]) -> List[List[str]]:
    rows: List[List[str]] = []
    headers = SHEET_NAME_TO_SPEC["Notes"].columns

    for entry in note_entries:
        scope = str(entry.get("scope") or "")
        row_key = str(entry.get("rowKey") or entry.get("row_key") or "")
        notes = entry.get("notes") or []
        if not scope or not row_key or not isinstance(notes, list):
            continue

        ordered = sorted(notes, key=lambda n: str(n.get("createdAt") or ""))
        seq = 0
        for note in ordered:
            if not isinstance(note, dict):
                continue
            seq += 1
            rows.append(
                [
                    scope,
                    row_key,
                    str(note.get("id") or ""),
                    "",
                    str(seq),
                    str(note.get("createdAt") or ""),
                    str(note.get("author") or ""),
                    "TRUE" if note.get("completed") else "FALSE",
                    str(note.get("note") or ""),
                    "existing",
                ]
            )
            replies = sorted(
                note.get("replies") or [],
                key=lambda r: str(r.get("createdAt") or "") if isinstance(r, dict) else "",
            )
            for reply in replies:
                if not isinstance(reply, dict):
                    continue
                seq += 1
                rows.append(
                    [
                        scope,
                        row_key,
                        str(reply.get("id") or ""),
                        str(note.get("id") or ""),
                        str(seq),
                        str(reply.get("createdAt") or ""),
                        str(reply.get("author") or ""),
                        "",
                        str(reply.get("note") or ""),
                        "existing",
                    ]
                )
    return rows


def apply_notes_to_workbook(workbook: Workbook, note_entries: List[Dict[str, Any]]) -> None:
    prev_by_scope_row: Dict[tuple[str, str], str] = {}
    for entry in note_entries:
        if not isinstance(entry, dict):
            continue
        scope = str(entry.get("scope") or "").strip()
        row_key = normalize_spreadsheet_string_cell(
            entry.get("rowKey") or entry.get("row_key") or ""
        )
        notes = entry.get("notes") or []
        if not scope or not row_key or not notes:
            continue
        prev_by_scope_row[(scope, row_key)] = format_prev_notes_display(notes)

    for sheet_name, scope, id_column in ENTITY_NOTE_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        values = {
            row_key: text
            for (entry_scope, row_key), text in prev_by_scope_row.items()
            if entry_scope == scope
        }
        _set_column_values_for_keys(ws, "prev_notes", values, id_column)
        _clear_column(ws, "new_notes", id_column)

    if "Notes" not in workbook.sheetnames:
        workbook.create_sheet("Notes")
    notes_ws = workbook["Notes"]
    headers = SHEET_NAME_TO_SPEC["Notes"].columns
    if notes_ws.max_row:
        notes_ws.delete_rows(1, notes_ws.max_row)
    notes_ws.append(headers)
    for row in _notes_sheet_rows(note_entries):
        notes_ws.append(row)
