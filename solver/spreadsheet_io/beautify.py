from __future__ import annotations

from typing import Dict, Iterable, List, Set

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from spreadsheet_io.spreadsheet_utils import SPREADSHEET_SPECS
except ModuleNotFoundError:
    from spreadsheet_utils import SPREADSHEET_SPECS  # type: ignore[no-redef]

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="E2E8F0")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

_GRID = Side(style="thin", color="525252")
_CELL_BORDER = Border(left=_GRID, right=_GRID, top=_GRID, bottom=_GRID)
_HEADER_BORDER = Border(
    left=_GRID,
    right=_GRID,
    top=_GRID,
    bottom=Side(style="medium", color="262626"),
)

PREV_NOTES_FILL = PatternFill(fill_type="solid", fgColor="E8E8E8")
NEW_NOTES_FILL = PatternFill(fill_type="solid", fgColor="FFF9E6")

# Columns that should wrap and get wider caps
DEFAULT_WRAP_COLUMNS: Set[str] = {
    "prev_notes",
    "new_notes",
    "body",
    "compatible_timeslot_sets",
    "allowed_meeting_patterns",
    "room_requirements",
    "unavailable_times",
    "member_section_ids",
    "reason",
    "tags",
}

WIDE_COLUMNS: Set[str] = {
    "prev_notes",
    "new_notes",
    "body",
    "compatible_timeslot_sets",
}

NARROW_COLUMNS: Set[str] = {
    "id",
    "state",
    "section_code",
    "day",
    "slots_required",
    "capacity",
    "weight",
    "seq",
    "completed",
    "source",
}


def _cell_display_length(value: object) -> int:
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0
    return max(len(line) for line in text.splitlines()) if "\n" in text else len(text)


def _column_width(header: str, max_content_len: int) -> float:
    if header in WIDE_COLUMNS:
        return float(min(max(max_content_len + 2, 28), 72))
    if header in NARROW_COLUMNS:
        return float(min(max(max_content_len + 2, 8), 18))
    return float(min(max(max_content_len + 2, 12), 42))


def beautify_worksheet(
    ws: Worksheet,
    headers: List[str],
    *,
    wrap_columns: Iterable[str] | None = None,
    column_fills: Dict[str, PatternFill] | None = None,
    sample_row_limit: int = 600,
) -> None:
    wrap_set = set(wrap_columns) if wrap_columns is not None else set(DEFAULT_WRAP_COLUMNS)
    fills = column_fills or {}
    max_row = ws.max_row or 1
    width_sample_last_row = min(max_row, sample_row_limit + 1)

    for col_idx, header in enumerate(headers, start=1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font = HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.alignment = HEADER_ALIGN
        header_cell.border = _HEADER_BORDER

        max_len = len(header)
        fill = fills.get(header)

        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _CELL_BORDER
            if header in wrap_set:
                cell.alignment = WRAP_TOP
            if fill is not None:
                cell.fill = fill

        for row_idx in range(2, width_sample_last_row + 1):
            max_len = max(
                max_len,
                _cell_display_length(ws.cell(row=row_idx, column=col_idx).value),
            )

        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width(header, max_len)

    ws.sheet_view.showGridLines = True

    if max_row > 1:
        ws.freeze_panes = "A2"


def beautify_workbook(workbook: Workbook) -> None:
    specs_by_name = {spec.name: spec for spec in SPREADSHEET_SPECS}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        spec = specs_by_name.get(sheet_name)
        if not spec:
            continue

        fills: Dict[str, PatternFill] = {}
        if sheet_name in {
            "Sections",
            "Instructors",
            "Rooms",
            "Timeslots",
            "MeetingPatterns",
        }:
            fills["prev_notes"] = PREV_NOTES_FILL
            fills["new_notes"] = NEW_NOTES_FILL
        if sheet_name == "Notes":
            fills["body"] = PREV_NOTES_FILL

        beautify_worksheet(ws, spec.columns, column_fills=fills)
