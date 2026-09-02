from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LIST_SEPARATOR = ";"
SET_SEPARATOR = "|"


@dataclass(frozen=True)
class SheetSpec:
    name: str
    columns: List[str]


SPREADSHEET_SPECS: List[SheetSpec] = [
    SheetSpec(
        name="Sections",
        columns=[
            "id",
            "course_id",
            "department",
            "section_code",
            "section_number",
            "instructor_id",
            "expected_enrollment",
            "enrollment_cap",
            "allowed_meeting_patterns",
            "room_requirements",
            "crosslist_group_id",
            "tags",
            "previous_meeting_pattern",
            "state",
            "duration",
            "prev_notes",
            "new_notes",
        ],
    ),
    SheetSpec(
        name="Instructors",
        columns=[
            "id",
            "name",
            "rank_type",
            "unavailable_times",
            "preferred_days",
            "preferred_patterns",
            "max_teaching_days",
            "prev_notes",
            "new_notes",
        ],
    ),
    SheetSpec(
        name="Rooms",
        columns=[
            "id",
            "building",
            "room_number",
            "capacity",
            "features",
            "prev_notes",
            "new_notes",
        ],
    ),
    SheetSpec(
        name="Timeslots",
        columns=[
            "id",
            "day",
            "start_time",
            "end_time",
            "slot_type",
            "prev_notes",
            "new_notes",
        ],
    ),
    SheetSpec(
        name="MeetingPatterns",
        columns=[
            "id",
            "slots_required",
            "allowed_days",
            "compatible_timeslot_sets",
            "prev_notes",
            "new_notes",
        ],
    ),
    SheetSpec(
        name="Notes",
        columns=[
            "scope",
            "row_key",
            "note_id",
            "parent_note_id",
            "seq",
            "created_at",
            "author",
            "completed",
            "body",
            "source",
        ],
    ),
    SheetSpec(
        name="CrosslistGroups",
        columns=["id", "member_section_ids"],
    ),
    SheetSpec(
        name="NoOverlapGroups",
        columns=["id", "member_section_ids", "reason"],
    ),
    SheetSpec(
        name="BlockedTimes",
        columns=[
            "scope",
            "days",
            "start_time",
            "end_time",
            "instructor_id",
            "room_id",
            "timeslot_ids",
            "reason",
        ],
    ),
    SheetSpec(
        name="LockedAssignments",
        columns=["section_id", "fixed_timeslot_set", "fixed_room"],
    ),
    SheetSpec(
        name="SoftLocks",
        columns=["section_id", "preferred_timeslot_set", "preferred_room", "weight"],
    ),
]

SHEET_NAME_TO_SPEC: Dict[str, SheetSpec] = {spec.name: spec for spec in SPREADSHEET_SPECS}

# Sections schema before `duration` was renamed from `semester_length`.
LEGACY_SHEET_COLUMNS_V4: Dict[str, List[str]] = {
    "Sections": [
        "id",
        "course_id",
        "department",
        "section_code",
        "section_number",
        "instructor_id",
        "expected_enrollment",
        "enrollment_cap",
        "allowed_meeting_patterns",
        "room_requirements",
        "crosslist_group_id",
        "tags",
        "previous_meeting_pattern",
        "state",
        "semester_length",
    ],
}

# Sections schema before `duration` / `semester_length` (includes section_number).
LEGACY_SHEET_COLUMNS_V3: Dict[str, List[str]] = {
    "Sections": [
        "id",
        "course_id",
        "department",
        "section_code",
        "section_number",
        "instructor_id",
        "expected_enrollment",
        "enrollment_cap",
        "allowed_meeting_patterns",
        "room_requirements",
        "crosslist_group_id",
        "tags",
        "previous_meeting_pattern",
        "state",
    ],
}

# Sections schema before the `section_number` column was introduced. Still accepted
# on import so spreadsheets exported by older versions of this app keep working.
LEGACY_SHEET_COLUMNS_V2: Dict[str, List[str]] = {
    "Sections": [
        "id",
        "course_id",
        "department",
        "section_code",
        "instructor_id",
        "expected_enrollment",
        "enrollment_cap",
        "allowed_meeting_patterns",
        "room_requirements",
        "crosslist_group_id",
        "tags",
        "previous_meeting_pattern",
        "state",
    ],
}

LEGACY_SHEET_COLUMNS: Dict[str, List[str]] = {
    "Sections": [
        "id",
        "course_id",
        "section_code",
        "instructor_id",
        "expected_enrollment",
        "enrollment_cap",
        "allowed_meeting_patterns",
        "room_requirements",
        "crosslist_group_id",
        "tags",
    ],
    "Instructors": [
        "id",
        "rank_type",
        "unavailable_times",
        "preferred_days",
        "preferred_patterns",
        "max_teaching_days",
    ],
    "Rooms": ["id", "building", "capacity", "features"],
    "Timeslots": ["id", "day", "start_time", "end_time"],
}


_NOTE_SUFFIX = ("prev_notes", "new_notes")
_ENTITY_SHEETS_WITH_NOTES = frozenset(
    {"Sections", "Instructors", "Rooms", "Timeslots", "MeetingPatterns"}
)


def _resolve_sections_headers(matched: List[str], canonical: List[str]) -> List[str]:
    """Map legacy Sections headers (e.g. semester_length) to canonical names."""
    if matched == canonical:
        return canonical
    if "semester_length" in matched and "duration" in canonical:
        legacy_with_duration = [
            "duration" if column == "semester_length" else column for column in matched
        ]
        if legacy_with_duration == canonical:
            return canonical
    return matched


def normalize_sheet_headers(sheet_name: str, headers: List[str]) -> List[str]:
    """
    Accept either the current schema or a legacy schema for certain sheets.

    Returns the canonical header list for the sheet if the provided headers are compatible.
    Note columns (prev_notes, new_notes) are never returned here — the solver ignores them;
    the platform reads those cells separately for note import/export.
    """
    spec = SHEET_NAME_TO_SPEC[sheet_name]
    expected = spec.columns

    candidate_schemas = [expected]
    for legacy_map in (LEGACY_SHEET_COLUMNS_V4, LEGACY_SHEET_COLUMNS_V3, LEGACY_SHEET_COLUMNS_V2):
        extra = legacy_map.get(sheet_name)
        if extra:
            candidate_schemas.append(extra)

    if sheet_name in _ENTITY_SHEETS_WITH_NOTES:
        canonical_scheduling = [c for c in expected if c not in _NOTE_SUFFIX]
        for schema in candidate_schemas:
            scheduling_only = [c for c in schema if c not in _NOTE_SUFFIX]
            with_notes = scheduling_only + list(_NOTE_SUFFIX)
            if headers[: len(with_notes)] == with_notes:
                if sheet_name == "Sections":
                    return _resolve_sections_headers(scheduling_only, canonical_scheduling)
                return scheduling_only
            if headers[: len(scheduling_only)] == scheduling_only:
                if sheet_name == "Sections":
                    return _resolve_sections_headers(scheduling_only, canonical_scheduling)
                return scheduling_only

    if headers[: len(expected)] == expected:
        return expected

    for legacy_map in (LEGACY_SHEET_COLUMNS_V4, LEGACY_SHEET_COLUMNS_V3, LEGACY_SHEET_COLUMNS_V2):
        legacy = legacy_map.get(sheet_name)
        if legacy and headers[: len(legacy)] == legacy:
            if sheet_name == "Sections":
                canonical_scheduling = [c for c in expected if c not in _NOTE_SUFFIX]
                return _resolve_sections_headers(legacy, canonical_scheduling)
            return legacy

    legacy = LEGACY_SHEET_COLUMNS.get(sheet_name)
    if legacy and headers[: len(legacy)] == legacy:
        return legacy

    raise ValueError(
        f"Sheet '{sheet_name}' has invalid headers. Expected: {expected}. Found: {headers[:len(expected)]}"
    )


def normalize_section_state(raw: Any) -> str:
    """Return 'active', 'new', or 'archived'. Blank/missing defaults to active."""
    value = str(raw or "").strip().lower()
    if value in ("archived", "archive"):
        return "archived"
    if value == "new":
        return "new"
    return "active"


SEMESTER_LENGTH_LABELS: Dict[str, str] = {
    "full": "Full",
    "half_any": "Half (any)",
    "first_half": "First Half",
    "second_half": "Second Half",
}

_SEMESTER_LENGTH_ALIASES: Dict[str, str] = {
    "full": "full",
    "half any": "half_any",
    "half (any)": "half_any",
    "half": "half_any",
    "first half": "first_half",
    "1st half": "first_half",
    "second half": "second_half",
    "2nd half": "second_half",
}


def normalize_semester_length(raw: Any) -> str:
    """Return full, half_any, first_half, or second_half. Blank/missing defaults to full."""
    value = re.sub(r"[_-]+", " ", str(raw or "").strip().lower())
    value = re.sub(r"\s+", " ", value)
    if not value:
        return "full"
    return _SEMESTER_LENGTH_ALIASES.get(value, "full")


def semester_length_label(raw: Any) -> str:
    return SEMESTER_LENGTH_LABELS[normalize_semester_length(raw)]


def apply_semester_length_dropdown(ws: Worksheet) -> None:
    """Excel list validation for Sections.duration (human-readable labels)."""
    spec = SHEET_NAME_TO_SPEC.get("Sections")
    if spec is None:
        return
    try:
        col_idx = spec.columns.index("duration") + 1
    except ValueError:
        return
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    col_letter = get_column_letter(col_idx)
    end_row = max(ws.max_row, 200)
    formula = ",".join(SEMESTER_LENGTH_LABELS.values())
    dv = DataValidation(
        type="list",
        formula1=f'"{formula}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Duration",
        error="Select Full, Half (any), First Half, or Second Half.",
    )
    dv.sqref = f"{col_letter}2:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def parse_list_cell(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split(LIST_SEPARATOR) if part.strip()]


def serialize_list_cell(values: List[str] | None) -> str:
    if not values:
        return ""
    return LIST_SEPARATOR.join(str(v).strip() for v in values if str(v).strip())


def _is_missing_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and value != value:  # NaN
        return True
    try:
        import pandas as pd

        if pd.isna(value):
            return True
    except Exception:
        pass
    return False


def parse_nested_list_cell(value: Any) -> List[List[str]]:
    """
    Parse MeetingPatterns.compatible_timeslot_sets from a spreadsheet cell.

    Canonical text form: alternatives separated by ';', timeslot IDs within one
    alternative separated by '|', e.g. "B1-M|B1-W;B2-M|B2-F".

    Also accepts list-of-lists (JSON / DB shape) and a flat list of IDs (one set).
    """
    if _is_missing_cell(value):
        return []
    if isinstance(value, (list, tuple)):
        if not value:
            return []
        if all(isinstance(x, (list, tuple)) for x in value):
            sets: List[List[str]] = []
            for inner in value:
                entries = [
                    str(x).strip()
                    for x in inner
                    if str(x).strip() and not _is_missing_cell(x)
                ]
                if entries:
                    sets.append(entries)
            return sets
        entries = [
            str(x).strip()
            for x in value
            if str(x).strip() and not _is_missing_cell(x)
        ]
        return [entries] if entries else []

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return []
    sets = []
    for set_part in text.split(LIST_SEPARATOR):
        set_part = set_part.strip()
        if not set_part:
            continue
        inner_sep = SET_SEPARATOR if SET_SEPARATOR in set_part else ","
        entries = [part.strip() for part in set_part.split(inner_sep) if part.strip()]
        if entries:
            sets.append(entries)
    return sets


def serialize_nested_list_cell(values: Any) -> str:
    """Write compatible_timeslot_sets as ';'-separated alternatives, '|' within a set."""
    if _is_missing_cell(values):
        return ""
    if isinstance(values, str):
        parsed = parse_nested_list_cell(values)
        return serialize_nested_list_cell(parsed)
    if not isinstance(values, list) or not values:
        return ""
    rows: List[List[str]]
    if isinstance(values[0], list):
        rows = values  # type: ignore[assignment]
    else:
        rows = [values]  # type: ignore[list-item]
    set_strings: List[str] = []
    for row in rows:
        if not isinstance(row, (list, tuple)):
            continue
        entries = [str(v).strip() for v in row if str(v).strip()]
        if entries:
            set_strings.append(SET_SEPARATOR.join(entries))
    return LIST_SEPARATOR.join(set_strings)


def parse_bool_cell(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def maybe_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def maybe_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def normalize_spreadsheet_string_cell(value: Any) -> str:
    """
    Canonical string form for IDs, codes, room numbers, and other text cells.

    Excel `data_only` and pandas often yield floats (e.g. 101.0); DB/TS use strings.
    """
    if value is None or _is_missing_cell(value):
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value == int(value):
            return str(int(value))
        return str(value)
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    # String that looks like a whole float, e.g. from JSON "101.0"
    if len(text) >= 3 and text[-2:] == ".0" and text[:-2].lstrip("-").isdigit():
        return text[:-2]
    return text


def _pad_digit_run(digits: str) -> str:
    """
    If fewer than 3 digits and does not already begin with 0, prepend one zero.
    Examples: 2 -> 02, 12 -> 012, 02 -> 02, 102 -> 102.
    Also collapses prior 3-digit zfill of single-digit rooms (002 -> 02).
    """
    if not digits.isdigit():
        return digits
    if re.fullmatch(r"00\d", digits):
        return digits[1:]
    if len(digits) >= 3 or digits.startswith("0"):
        return digits
    return f"0{digits}"


def canonicalize_room_number(value: Any) -> str:
    """
    Canonical room_number string.

    Short numbers without a leading zero get one prepended (2 -> 02; 02 stays).
    Mixed values pad the final numeric run only (e.g. A2 -> A02). Idempotent.
    """
    text = normalize_spreadsheet_string_cell(value)
    if not text:
        return ""
    if text.isdigit():
        return _pad_digit_run(text)
    match = re.match(r"^(.*?)(\d+)(\D*)$", text)
    if not match:
        return text
    prefix, digits, suffix = match.group(1), match.group(2), match.group(3)
    return f"{prefix}{_pad_digit_run(digits)}{suffix}"


def maybe_str(value: Any) -> str | None:
    """Optional string from a cell; normalizes Excel/JSON float artifacts."""
    if value is None:
        return None
    text = normalize_spreadsheet_string_cell(value)
    return text or None


def format_room_number_for_export(value: Any) -> str:
    """Export room_number as canonical string."""
    return canonicalize_room_number(value)


def build_template_workbook() -> Workbook:
    try:
        from spreadsheet_io.beautify import beautify_workbook
    except ModuleNotFoundError:
        from beautify import beautify_workbook  # type: ignore[no-redef]

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for spec in SPREADSHEET_SPECS:
        ws = wb.create_sheet(spec.name)
        ws.append(spec.columns)
        if spec.name == "Sections":
            apply_semester_length_dropdown(ws)

    beautify_workbook(wb)
    return wb


def build_template_bytes() -> bytes:
    wb = build_template_workbook()
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _autosize_columns(ws: Worksheet, headers: List[str]) -> None:
    """Legacy helper; prefer :func:`beautify.beautify_workbook`."""
    from openpyxl.utils import get_column_letter

    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(len(header) + 2, 14)
