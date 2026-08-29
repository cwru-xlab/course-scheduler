"""Build read-only Online Schedule rows for spreadsheet export."""

from __future__ import annotations

import re
from typing import Any, Dict, List, Optional, Tuple

try:
    from online_sections import ONLINE_ROOM_SENTINEL, is_online_section
    from spreadsheet_io.spreadsheet_utils import normalize_section_state
except ModuleNotFoundError:
    from online_sections import ONLINE_ROOM_SENTINEL, is_online_section  # type: ignore[no-redef]
    from spreadsheet_utils import normalize_section_state  # type: ignore[no-redef]

WEEKDAY_ORDER = ("Mon", "Tue", "Wed", "Thu", "Fri")

ONLINE_SCHEDULE_COLUMNS = (
    "Day",
    "Start",
    "End",
    "Department",
    "Course",
    "Section",
    "Instructor",
    "Section ID",
)

_DAY_NAME_TO_SHORT = {
    "mon": "Mon",
    "monday": "Mon",
    "m": "Mon",
    "tue": "Tue",
    "tues": "Tue",
    "tuesday": "Tue",
    "tu": "Tue",
    "t": "Tue",
    "wed": "Wed",
    "weds": "Wed",
    "wednesday": "Wed",
    "w": "Wed",
    "thu": "Thu",
    "thur": "Thu",
    "thurs": "Thu",
    "thursday": "Thu",
    "th": "Thu",
    "r": "Thu",
    "fri": "Fri",
    "friday": "Fri",
    "f": "Fri",
}


def _normalize_weekday_token(token: str) -> Optional[str]:
    key = token.strip().lower()
    if not key:
        return None
    day = _DAY_NAME_TO_SHORT.get(key)
    if day in WEEKDAY_ORDER:
        return day
    return None


def expand_timeslot_weekdays(timeslot: Dict[str, Any]) -> List[str]:
    """Return Mon–Fri short day codes represented by a timeslot."""
    raw = str(timeslot.get("days") or timeslot.get("day") or "").strip()
    if not raw:
        return []

    single = _normalize_weekday_token(raw)
    if single:
        return [single]

    tokens = re.split(r"[^A-Za-z]+", raw)
    days: List[str] = []
    for token in tokens:
        day = _normalize_weekday_token(token)
        if day and day not in days:
            days.append(day)

    compact = re.sub(r"[^A-Za-z]", "", raw).upper()
    if compact and not days:
        idx = 0
        while idx < len(compact):
            two = compact[idx : idx + 2]
            if two == "TU":
                if "Tue" not in days:
                    days.append("Tue")
                idx += 2
                continue
            if two == "TH":
                if "Thu" not in days:
                    days.append("Thu")
                idx += 2
                continue
            letter = compact[idx]
            mapped = {
                "M": "Mon",
                "W": "Wed",
                "F": "Fri",
                "R": "Thu",
                "T": "Tue",
            }.get(letter)
            if mapped and mapped not in days:
                days.append(mapped)
            idx += 1

    return [day for day in days if day in WEEKDAY_ORDER]


def _timeslot_lookup(payload: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    for slot in payload.get("timeslots") or []:
        if not isinstance(slot, dict):
            continue
        slot_id = str(slot.get("id") or "").strip()
        if slot_id:
            lookup[slot_id] = slot
    return lookup


def _instructor_lookup(payload: Dict[str, Any]) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for inst in payload.get("instructors") or []:
        if not isinstance(inst, dict):
            continue
        inst_id = str(inst.get("id") or "").strip()
        if not inst_id:
            continue
        lookup[inst_id] = str(inst.get("name") or inst_id).strip() or inst_id
    return lookup


def _resolve_timeslot_ids(section_id: str, section: Dict[str, Any], payload: Dict[str, Any]) -> List[str]:
    resolved: List[str] = []
    seen: set[str] = set()

    for lock in payload.get("locked_assignments") or []:
        if not isinstance(lock, dict):
            continue
        if str(lock.get("section_id") or "").strip() != section_id:
            continue
        fixed_room = str(lock.get("fixed_room") or "").strip()
        if fixed_room and fixed_room not in ("", ONLINE_ROOM_SENTINEL):
            continue
        for ts_id in lock.get("fixed_timeslot_set") or []:
            token = str(ts_id).strip()
            if token and token not in seen:
                seen.add(token)
                resolved.append(token)

    if resolved:
        return resolved

    raw_ids = section.get("timeslot_ids")
    if isinstance(raw_ids, list):
        ids = [str(ts).strip() for ts in raw_ids if str(ts).strip()]
        if ids:
            return ids

    section_ts = str(section.get("timeslot_id") or "").strip()
    if section_ts:
        return [section_ts]

    solution = payload.get("solution")
    if isinstance(solution, dict):
        for assignment in solution.get("assignments") or []:
            if not isinstance(assignment, dict):
                continue
            if str(assignment.get("section_id") or "").strip() != section_id:
                continue
            room_id = str(assignment.get("room_id") or "").strip()
            if room_id and room_id != ONLINE_ROOM_SENTINEL:
                continue
            ids = [
                str(ts).strip()
                for ts in (assignment.get("timeslot_ids") or [])
                if str(ts).strip()
            ]
            if ids:
                return ids

    return []


def build_online_schedule_rows(payload: Dict[str, Any]) -> List[Dict[str, str]]:
    """Return flat data rows (no day headers) for scheduled online sections."""
    timeslot_by_id = _timeslot_lookup(payload)
    instructor_by_id = _instructor_lookup(payload)
    rows: List[Dict[str, str]] = []

    for section in payload.get("sections") or []:
        if not isinstance(section, dict):
            continue
        if normalize_section_state(section.get("state")) == "archived":
            continue
        if not is_online_section(section):
            continue

        section_id = str(section.get("id") or "").strip()
        if not section_id:
            continue

        timeslot_ids = _resolve_timeslot_ids(section_id, section, payload)
        if not timeslot_ids:
            continue

        department = str(section.get("department") or "").strip()
        course = str(section.get("course_id") or "").strip()
        section_code = str(section.get("section_code") or "").strip()
        section_number = str(section.get("section_number") or "").strip()
        instructor_id = str(section.get("instructor_id") or "").strip()
        instructor = instructor_by_id.get(instructor_id, instructor_id)

        for ts_id in timeslot_ids:
            slot = timeslot_by_id.get(ts_id)
            if not slot:
                continue
            start = str(slot.get("start_time") or "").strip()
            end = str(slot.get("end_time") or "").strip()
            for day in expand_timeslot_weekdays(slot):
                rows.append(
                    {
                        "Day": day,
                        "Start": start,
                        "End": end,
                        "Department": department,
                        "Course": course,
                        "Section": section_number or section_code,
                        "Instructor": instructor,
                        "Section ID": section_id,
                    }
                )

    rows.sort(
        key=lambda row: (
            WEEKDAY_ORDER.index(row["Day"]) if row["Day"] in WEEKDAY_ORDER else 99,
            row["Start"],
            row["Department"],
            row["Course"],
            row["Section ID"],
        )
    )
    return rows


def build_online_schedule_sheet_rows(payload: Dict[str, Any]) -> List[Tuple[str, ...]]:
    """Return worksheet rows including subtitle, headers, and day-group labels."""
    data_rows = build_online_schedule_rows(payload)
    sheet_rows: List[Tuple[str, ...]] = [
        (
            "Read-only summary. Edit sections and locks on entity sheets.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
        ONLINE_SCHEDULE_COLUMNS,
    ]

    current_day: Optional[str] = None
    for row in data_rows:
        day = row["Day"]
        if day != current_day:
            current_day = day
            sheet_rows.append((day, "", "", "", "", "", "", ""))
        sheet_rows.append(
            (
                "",
                row["Start"],
                row["End"],
                row["Department"],
                row["Course"],
                row["Section"],
                row["Instructor"],
                row["Section ID"],
            )
        )

    if len(sheet_rows) == 2:
        sheet_rows.append(
            ("No scheduled online sections.", "", "", "", "", "", "", ""),
        )

    return sheet_rows


def apply_online_schedule_sheet(
    workbook: Any,
    payload: Dict[str, Any],
) -> None:
    """Add export-only Online Schedule worksheet."""
    from openpyxl.styles import Font

    ws = workbook.create_sheet("Online Schedule")
    for row in build_online_schedule_sheet_rows(payload):
        ws.append(list(row))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ONLINE_SCHEDULE_COLUMNS))
    subtitle = ws.cell(row=1, column=1)
    subtitle.font = Font(italic=True, color="FF475569")

    header_row_idx = 2
    for col_idx in range(1, len(ONLINE_SCHEDULE_COLUMNS) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = Font(bold=True)

    for row_idx in range(3, ws.max_row + 1):
        day_cell = ws.cell(row=row_idx, column=1)
        if day_cell.value and not ws.cell(row=row_idx, column=2).value:
            day_cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 16
