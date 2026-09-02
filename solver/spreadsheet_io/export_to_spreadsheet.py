from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from spreadsheet_io.beautify import beautify_workbook
    from spreadsheet_io.notes_export import apply_notes_to_workbook
    from spreadsheet_io.online_schedule_export import apply_online_schedule_sheet
    from spreadsheet_io.spreadsheet_utils import (
        SPREADSHEET_SPECS,
        canonicalize_room_number,
        format_room_number_for_export,
        normalize_spreadsheet_string_cell,
        serialize_list_cell,
        serialize_nested_list_cell,
        apply_semester_length_dropdown,
        semester_length_label,
    )
except ModuleNotFoundError:
    from beautify import beautify_workbook  # type: ignore[no-redef]
    from notes_export import apply_notes_to_workbook  # type: ignore[no-redef]
    from online_schedule_export import apply_online_schedule_sheet  # type: ignore[no-redef]
    from spreadsheet_utils import (  # type: ignore[no-redef]
        SPREADSHEET_SPECS,
        canonicalize_room_number,
        format_room_number_for_export,
        normalize_spreadsheet_string_cell,
        serialize_list_cell,
        serialize_nested_list_cell,
        apply_semester_length_dropdown,
        semester_length_label,
    )


def _export_str(value: Any) -> str:
    """Normalize string/ID cells so Excel does not show spurious `.0`."""
    return normalize_spreadsheet_string_cell(value)


def _build_instructor_name_lookup(
    payload: Dict[str, Any],
) -> Dict[str, str]:
    """Build a mapping of instructor_id → display name."""
    lookup: Dict[str, str] = {}
    for inst in payload.get("instructors", []):
        inst_id = str(inst.get("id", "")).strip()
        name = str(inst.get("name") or inst.get("id", "")).strip()
        if inst_id:
            lookup[inst_id] = name or inst_id
    return lookup


def _instructor_name(
    instructor_id: str,
    lookup: Dict[str, str],
) -> str:
    """Return the display name for an instructor_id, or the raw id if not found."""
    return lookup.get(instructor_id, instructor_id)


def scheduling_input_to_excel_bytes(
    payload: Dict[str, Any],
    note_entries: List[Dict[str, Any]] | None = None,
) -> bytes:
    workbook = Workbook()
    default_ws = workbook.active
    workbook.remove(default_ws)

    instructor_lookup = _build_instructor_name_lookup(payload)
    instructor_names = sorted(set(instructor_lookup.values()))

    sheets: Dict[str, Any] = {}
    for spec in SPREADSHEET_SPECS:
        ws = workbook.create_sheet(spec.name)
        sheets[spec.name] = ws
        ws.append(spec.columns)
        rows = _rows_for_sheet(spec.name, payload, instructor_lookup)
        for row in rows:
            ws.append([row.get(column, "") for column in spec.columns])

    # Force Rooms.room_number cells to Excel text so leading zeros survive re-import.
    rooms_ws = sheets.get("Rooms")
    if rooms_ws is not None:
        try:
            room_number_col = SPREADSHEET_SPECS[
                next(i for i, s in enumerate(SPREADSHEET_SPECS) if s.name == "Rooms")
            ].columns.index("room_number") + 1
        except (StopIteration, ValueError):
            room_number_col = None
        if room_number_col is not None:
            for row_idx in range(2, rooms_ws.max_row + 1):
                cell = rooms_ws.cell(row=row_idx, column=room_number_col)
                if cell.value is None or cell.value == "":
                    continue
                cell.value = canonicalize_room_number(cell.value)
                cell.number_format = "@"

    instructor_ws = sheets.get("Instructors")
    for spec in SPREADSHEET_SPECS:
        if instructor_names and spec.name in ("Sections", "BlockedTimes"):
            _apply_instructor_dropdown(
                sheets[spec.name], spec.columns, instructor_names, instructor_ws,
            )
        if spec.name == "Sections":
            apply_semester_length_dropdown(sheets[spec.name])

    populated_notes = [
        entry
        for entry in (note_entries or [])
        if isinstance(entry, dict) and entry.get("notes")
    ]
    if populated_notes:
        apply_notes_to_workbook(workbook, populated_notes)

    apply_online_schedule_sheet(workbook, payload)

    beautify_workbook(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _apply_instructor_dropdown(
    ws: Any,
    columns: List[str],
    names: List[str],
    instructor_ws: Any = None,
) -> None:
    """Add a dropdown data validation to the instructor_id column.

    If *instructor_ws* is provided, the dropdown references the Name column
    on that sheet (avoids comma-in-name splitting in the formula string).
    """
    try:
        col_idx = columns.index("instructor_id") + 1  # 1-based
    except ValueError:
        return
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(col_idx)
    max_row = ws.max_row
    if max_row < 2:
        return

    if instructor_ws is not None and names:
        # Write names into a hidden helper column on the Instructors sheet
        # so the dropdown references cells (handles commas in names).
        helper_col_letter = get_column_letter(instructor_ws.max_column + 1)
        for i, name in enumerate(names, start=1):
            instructor_ws[f"{helper_col_letter}{i}"] = name
        formula = f"={instructor_ws.title}!${helper_col_letter}$1:${helper_col_letter}${len(names)}"
    else:
        formula = ",".join(names)

    dv = DataValidation(
        type="list",
        formula1=f'"{formula}"' if not formula.startswith("=") else formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Instructor",
        error="Select an instructor from the dropdown list.",
    )
    dv.sqref = f"{col_letter}2:{col_letter}{max_row}"
    ws.add_data_validation(dv)


def _rows_for_sheet(
    sheet_name: str,
    payload: Dict[str, Any],
    instructor_lookup: Dict[str, str] | None = None,
) -> list[Dict[str, Any]]:
    if sheet_name == "Sections":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "course_id": _export_str(item.get("course_id", "")),
                "department": _export_str(item.get("department") or ""),
                "section_code": _export_str(item.get("section_code", "")),
                "section_number": _export_str(item.get("section_number", "")),
                "instructor_id": _instructor_name(
                    _export_str(item.get("instructor_id", "")),
                    instructor_lookup or {},
                ),
                "expected_enrollment": item.get("expected_enrollment", ""),
                "enrollment_cap": item.get("enrollment_cap", ""),
                "allowed_meeting_patterns": serialize_list_cell(
                    item.get("allowed_meeting_patterns")
                ),
                "room_requirements": serialize_list_cell(item.get("room_requirements")),
                "crosslist_group_id": _export_str(item.get("crosslist_group_id") or ""),
                "tags": serialize_list_cell(item.get("tags")),
                "previous_meeting_pattern": _export_str(
                    item.get("previous_meeting_pattern") or ""
                ),
                "state": _export_str(item.get("state") or "active"),
                "duration": semester_length_label(item.get("semester_length")),
                "assigned_half": (
                    semester_length_label(item.get("assigned_half"))
                    if item.get("semester_length") == "half_any" and item.get("assigned_half")
                    else ""
                ),
                "prev_notes": "",
                "new_notes": "",
            }
            for item in payload.get("sections", [])
        ]
    if sheet_name == "Instructors":
        rows = []
        for item in payload.get("instructors", []):
            preferences = item.get("preferences") or {}
            rows.append(
                {
                    "id": _export_str(item.get("id", "")),
                    "name": _export_str(item.get("name") or ""),
                    "rank_type": _export_str(item.get("rank_type", "")),
                    "unavailable_times": serialize_list_cell(item.get("unavailable_times")),
                    "preferred_days": serialize_list_cell(preferences.get("preferred_days")),
                    "preferred_patterns": serialize_list_cell(
                        preferences.get("preferred_patterns")
                    ),
                    "max_teaching_days": _export_str(
                        preferences.get("max_teaching_days") or ""
                    ),
                    "prev_notes": "",
                    "new_notes": "",
                }
            )
        return rows
    if sheet_name == "Rooms":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "building": _export_str(item.get("building", "")),
                "room_number": format_room_number_for_export(item.get("room_number", "")),
                "capacity": item.get("capacity", ""),
                "features": serialize_list_cell(item.get("features")),
                "prev_notes": "",
                "new_notes": "",
            }
            for item in payload.get("rooms", [])
        ]
    if sheet_name == "Timeslots":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "day": _export_str(item.get("day", "")),
                "start_time": _export_str(item.get("start_time", "")),
                "end_time": _export_str(item.get("end_time", "")),
                "slot_type": _export_str(item.get("slot_type", "") or ""),
                "prev_notes": "",
                "new_notes": "",
            }
            for item in payload.get("timeslots", [])
        ]
    if sheet_name == "MeetingPatterns":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "slots_required": item.get("slots_required", ""),
                "allowed_days": serialize_list_cell(item.get("allowed_days")),
                "compatible_timeslot_sets": serialize_nested_list_cell(
                    item.get("compatible_timeslot_sets")
                ),
                "prev_notes": "",
                "new_notes": "",
            }
            for item in payload.get("meeting_patterns", [])
        ]
    if sheet_name == "CrosslistGroups":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "member_section_ids": serialize_list_cell(item.get("member_section_ids")),
            }
            for item in payload.get("crosslist_groups", [])
        ]
    if sheet_name == "NoOverlapGroups":
        return [
            {
                "id": _export_str(item.get("id", "")),
                "member_section_ids": serialize_list_cell(item.get("member_section_ids")),
                "reason": _export_str(item.get("reason", "")),
            }
            for item in payload.get("no_overlap_groups", [])
        ]
    if sheet_name == "BlockedTimes":
        return [
            {
                "scope": _export_str(item.get("scope", "")),
                "days": _export_str(item.get("days", "")),
                "start_time": _export_str(item.get("start_time", "")),
                "end_time": _export_str(item.get("end_time", "")),
                "instructor_id": _instructor_name(
                    _export_str(item.get("instructor_id", "")),
                    instructor_lookup or {},
                ),
                "room_id": _export_str(item.get("room_id", "")),
                "timeslot_ids": serialize_list_cell(item.get("timeslot_ids")),
                "reason": _export_str(item.get("reason", "")),
            }
            for item in payload.get("blocked_times", [])
        ]
    if sheet_name == "LockedAssignments":
        return [
            {
                "section_id": _export_str(item.get("section_id", "")),
                "fixed_timeslot_set": serialize_list_cell(item.get("fixed_timeslot_set")),
                "fixed_room": _export_str(item.get("fixed_room") or ""),
            }
            for item in payload.get("locked_assignments", [])
        ]
    if sheet_name == "SoftLocks":
        return [
            {
                "section_id": _export_str(item.get("section_id", "")),
                "preferred_timeslot_set": serialize_list_cell(item.get("preferred_timeslot_set")),
                "preferred_room": _export_str(item.get("preferred_room") or ""),
                "weight": item.get("weight", 1.0),
            }
            for item in payload.get("soft_locks", [])
        ]
    return []
