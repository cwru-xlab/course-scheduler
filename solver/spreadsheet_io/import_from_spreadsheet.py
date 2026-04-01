from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import load_workbook

import datetime

DEFAULT_ALLOWED_MEETING_PATTERNS = ["MP-A-50", "MP-B-75"]

try:
    from spreadsheet_io.spreadsheet_utils import (
        SHEET_NAME_TO_SPEC,
        format_room_number_for_export,
        maybe_float,
        maybe_int,
        maybe_str,
        normalize_sheet_headers,
        parse_bool_cell,
        parse_list_cell,
        parse_nested_list_cell,
    )
except ModuleNotFoundError:
    from spreadsheet_utils import (  # type: ignore[no-redef]
        SHEET_NAME_TO_SPEC,
        format_room_number_for_export,
        maybe_float,
        maybe_int,
        maybe_str,
        normalize_sheet_headers,
        parse_bool_cell,
        parse_list_cell,
        parse_nested_list_cell,
    )


def parse_scheduling_input_from_excel_bytes(excel_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)

    sections_rows = _read_rows(wb, "Sections")
    instructors_rows = _read_rows(wb, "Instructors")
    rooms_rows = _read_rows(wb, "Rooms")
    timeslots_rows = _read_rows(wb, "Timeslots")
    meeting_patterns_rows = _read_rows(wb, "MeetingPatterns")
    crosslist_rows = _read_rows(wb, "CrosslistGroups")
    no_overlap_rows = _read_rows(wb, "NoOverlapGroups")
    blocked_rows = _read_rows(wb, "BlockedTimes")
    locked_rows = _read_rows(wb, "LockedAssignments")
    soft_rows = _read_rows(wb, "SoftLocks")

    for row in timeslots_rows:
        for col in ["start_time", "end_time"]:
            val = row.get(col)
            if isinstance(val, datetime.time):
                row[col] = val.strftime("%H:%M")
            elif isinstance(val, str) and ":" in val:
                row[col] = val.strip()

    sections: List[Dict[str, Any]] = []
    seen_section_ids = set()
    for row in sections_rows:
        section_id = _required_str(row, "id", "Sections")
        if section_id in seen_section_ids:
            continue
        seen_section_ids.add(section_id)
        sections.append(
            {
                "id": section_id,
                "course_id": _str_with_default(row, "course_id", "Sections", default=""),
                "department": maybe_str(row.get("department")),
                "section_code": _str_with_default(row, "section_code", "Sections", default=""),
                "instructor_id": _str_with_default(row, "instructor_id", "Sections", default=""),
                "expected_enrollment": _int_with_default(
                    row, "expected_enrollment", "Sections", default=0
                ),
                "enrollment_cap": _int_with_default(row, "enrollment_cap", "Sections", default=0),
                "allowed_meeting_patterns": parse_list_cell(row.get("allowed_meeting_patterns")),
                "room_requirements": parse_list_cell(row.get("room_requirements")),
                "crosslist_group_id": maybe_str(row.get("crosslist_group_id")),
                "tags": parse_list_cell(row.get("tags")),
                "previous_meeting_pattern": maybe_str(row.get("previous_meeting_pattern")),
            }
        )

    instructors: List[Dict[str, Any]] = []
    for row in instructors_rows:
        instructor_id = _required_str(row, "id", "Instructors")
        instructors.append(
            {
                "id": instructor_id,
                "name": maybe_str(row.get("name")) or instructor_id,
                "rank_type": _str_with_default(row, "rank_type", "Instructors", default="Adjunct"),
                "unavailable_times": parse_list_cell(row.get("unavailable_times")),
                "preferences": {
                    "preferred_days": parse_list_cell(row.get("preferred_days")),
                    "preferred_patterns": parse_list_cell(row.get("preferred_patterns")),
                    "max_teaching_days": maybe_int(row.get("max_teaching_days")),
                },
            }
        )

    rooms: List[Dict[str, Any]] = []
    for row in rooms_rows:
        room_id = _required_str(row, "id", "Rooms")
        rooms.append(
            {
                "id": room_id,
                "building": _str_with_default(row, "building", "Rooms", default=""),
                "room_number": format_room_number_for_export(row.get("room_number")),
                "capacity": _int_with_default(row, "capacity", "Rooms", default=0),
                "features": parse_list_cell(row.get("features")),
            }
        )

    timeslots: List[Dict[str, Any]] = []
    for row in timeslots_rows:
        slot_id = _required_str(row, "id", "Timeslots")
        timeslots.append(
            {
                "id": slot_id,
                "day": _str_with_default(row, "day", "Timeslots", default=""),
                "start_time": _str_with_default(row, "start_time", "Timeslots", default=""),
                "end_time": _str_with_default(row, "end_time", "Timeslots", default=""),
                "slot_type": maybe_str(row.get("slot_type")),
            }
        )

    meeting_patterns: List[Dict[str, Any]] = []
    for row in meeting_patterns_rows:
        pattern_id = _required_str(row, "id", "MeetingPatterns")
        meeting_patterns.append(
            {
                "id": pattern_id,
                "slots_required": _int_with_default(
                    row, "slots_required", "MeetingPatterns", default=1
                ),
                "allowed_days": parse_list_cell(row.get("allowed_days")),
                "compatible_timeslot_sets": parse_nested_list_cell(
                    row.get("compatible_timeslot_sets")
                ),
            }
        )

    crosslist_groups: List[Dict[str, Any]] = []
    for row in crosslist_rows:
        crosslist_groups.append(
            {
                "id": _required_str(row, "id", "CrosslistGroups"),
                "member_section_ids": parse_list_cell(row.get("member_section_ids")),
                "require_same_room": parse_bool_cell(row.get("require_same_room"), default=False),
            }
        )

    no_overlap_groups: List[Dict[str, Any]] = []
    for row in no_overlap_rows:
        no_overlap_groups.append(
            {
                "id": _required_str(row, "id", "NoOverlapGroups"),
                "member_section_ids": parse_list_cell(row.get("member_section_ids")),
                "reason": _str_with_default(row, "reason", "NoOverlapGroups", default="constraint"),
            }
        )

    blocked_times: List[Dict[str, Any]] = []
    for row in blocked_rows:
        blocked_times.append(
            {
                "scope": _str_with_default(row, "scope", "BlockedTimes", default="global"),
                "timeslot_ids": parse_list_cell(row.get("timeslot_ids")),
                "reason": _str_with_default(row, "reason", "BlockedTimes", default="blocked"),
            }
        )

    locked_assignments: List[Dict[str, Any]] = []
    for row in locked_rows:
        lock: Dict[str, Any] = {"section_id": _required_str(row, "section_id", "LockedAssignments")}
        fixed_timeslot_set = parse_list_cell(row.get("fixed_timeslot_set"))
        fixed_room = maybe_str(row.get("fixed_room"))
        if fixed_timeslot_set:
            lock["fixed_timeslot_set"] = fixed_timeslot_set
        if fixed_room:
            lock["fixed_room"] = fixed_room
        locked_assignments.append(lock)

    soft_locks: List[Dict[str, Any]] = []
    for row in soft_rows:
        lock: Dict[str, Any] = {"section_id": _required_str(row, "section_id", "SoftLocks")}
        preferred_timeslot_set = parse_list_cell(row.get("preferred_timeslot_set"))
        preferred_room = maybe_str(row.get("preferred_room"))
        weight = maybe_float(row.get("weight"))
        lock["weight"] = weight if weight is not None else 1.0
        if preferred_timeslot_set:
            lock["preferred_timeslot_set"] = preferred_timeslot_set
        if preferred_room:
            lock["preferred_room"] = preferred_room
        soft_locks.append(lock)

    return {
        "sections": sections,
        "instructors": instructors,
        "rooms": rooms,
        "timeslots": timeslots,
        "meeting_patterns": meeting_patterns,
        "crosslist_groups": crosslist_groups,
        "no_overlap_groups": no_overlap_groups,
        "blocked_times": blocked_times,
        "locked_assignments": locked_assignments,
        "soft_locks": soft_locks,
    }


def _read_rows(workbook, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet: {sheet_name}")

    ws = workbook[sheet_name]
    
    header_row_index = 1
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if "id" in [str(v).lower().strip() if v else "" for v in row]:
            header_cells = row
            break
        header_row_index += 1
    else:
        # Fallback to row 1 if 'id' isn't found in the first 5 rows
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_row_index = 1

    if header_cells is None:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    headers = [str(v).strip() if v is not None else "" for v in header_cells]
    active_columns = normalize_sheet_headers(sheet_name, headers)

    rows: List[Dict[str, Any]] = []
    # Start reading from the row immediately AFTER the detected header
    for values in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if values is None:
            continue
        active_values = values[: len(active_columns)]
        if all(v is None or str(v).strip() == "" for v in active_values):
            continue
        row = {
            column: active_values[idx] if idx < len(active_values) else None
            for idx, column in enumerate(active_columns)
        }
        rows.append(row)
    return rows


def _required_str(row: Dict[str, Any], key: str, sheet: str) -> str:
    value = maybe_str(row.get(key))
    if value is None:
        raise ValueError(f"Sheet '{sheet}' missing required value for '{key}'.")
    return value


def _int_with_default(row: Dict[str, Any], key: str, sheet: str, default: int) -> int:
    value = maybe_int(row.get(key))
    return default if value is None else value


def _str_with_default(row: Dict[str, Any], key: str, sheet: str, default: str) -> str:
    value = maybe_str(row.get(key))
    return default if value is None else value
