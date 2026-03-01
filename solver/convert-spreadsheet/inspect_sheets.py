"""
Inspect xlsx files in input/ to discover sheet names, column headers, and format type.
Run from convert-spreadsheet directory: python inspect_sheets.py

Detects whether a file is in SIS format or SOC Editors format and reports
format-specific diagnostics (enrollment sums, cross-listing signals, session types, etc.)
"""
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

INPUT_DIR = Path(__file__).resolve().parent / "input"

# SOC Editors canonical column set (must match convert.py)
_SOC_EDITORS_COLUMNS = {
    "Term", "Subject", "Catalog", "Description", "Section", "Session",
    "Meeting Pattern Number", "Previous Semester Days and Times",
    "Meeting Dates", "Instructor(s)", "Enrl Cap (Cmbnd Enrl Cap)", "Instruction Mode",
}


def _detect_format(headers: list[str]) -> str:
    found = set(headers)
    overlap = found & _SOC_EDITORS_COLUMNS
    return "soc_editors" if len(overlap) >= 6 else "sis"


def _extract_class_nbr(section_val) -> str | None:
    if section_val is None:
        return None
    m = re.search(r"\((\d+)\)", str(section_val).strip())
    return m.group(1) if m else None


def _parse_enrollment_str(val) -> tuple[int, int]:
    """'21 (45)' -> (21, 45)."""
    if val is None:
        return 0, 0
    s = str(val).strip()
    m = re.match(r"^(\d+)\s*\((\d+)\)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = re.match(r"^(\d+)", s)
    if m2:
        v = int(m2.group(1))
        return v, v
    return 0, 0


def _inspect_soc_editors(rows: list[tuple], headers: list[str]) -> None:
    """Print SOC Editors-specific diagnostics."""
    h = {name: idx for idx, name in enumerate(headers)}

    def col(row, name):
        idx = h.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    data = rows  # already filtered to data rows by caller

    # -- Deduplication by class number (multi-pattern rows) --
    class_rows: dict[str, list] = defaultdict(list)
    missing_nbr = 0
    for row in data:
        nbr = _extract_class_nbr(col(row, "Section"))
        if nbr:
            class_rows[nbr].append(row)
        else:
            missing_nbr += 1

    unique_sections = len(class_rows)
    print(f"    Raw data rows:            {len(data)}")
    print(f"    Unique class numbers:     {unique_sections}")
    print(f"    Rows missing class nbr:   {missing_nbr}")
    print(f"    Multi-pattern sections:   "
          f"{sum(1 for v in class_rows.values() if len(v) > 1)} "
          f"(max occurrences: {max((len(v) for v in class_rows.values()), default=0)})")

    # -- Term --
    terms = set(str(col(r, "Term") or "").strip() for r in data)
    print(f"    Term(s):                  {', '.join(sorted(terms))}")

    # -- Subject distribution --
    subjects = Counter(str(col(r, "Subject") or "").strip() for r in data)
    print(f"    Unique subjects:          {len(subjects)}")
    print(f"    Subject distribution:     {dict(subjects.most_common(5))} {'...' if len(subjects) > 5 else ''}")

    # -- Session types --
    sessions = Counter(str(col(r, "Session") or "").strip() for r in data)
    print(f"    Session types:            {dict(sessions)}")

    # -- Instruction modes --
    modes = Counter()
    for r in data:
        m_str = str(col(r, "Instruction Mode") or "")
        m = re.match(r"^(\w+)", m_str)
        if m:
            modes[m.group(1)] += 1
    print(f"    Instruction modes:        {dict(modes)}")

    # -- Section types (LEC, SEM, IND, etc.) --
    sec_types = Counter()
    for r in data:
        sec_val = str(col(r, "Section") or "")
        m = re.search(r"\d+-(\w+)", sec_val)
        if m:
            sec_types[m.group(1)] += 1
    print(f"    Section types:            {dict(sec_types)}")

    # -- Enrollment sums (using first row per class number) --
    total_sec_cap = 0
    total_cmbnd_cap = 0
    for rlist in class_rows.values():
        sec_cap, cmbnd_cap = _parse_enrollment_str(col(rlist[0], "Enrl Cap (Cmbnd Enrl Cap)"))
        total_sec_cap += sec_cap
        total_cmbnd_cap += cmbnd_cap
    print(f"    Total section cap:        {total_sec_cap}")
    print(f"    Total combined cap:       {total_cmbnd_cap}")

    # -- Instructors --
    instructors = set()
    staff_count = 0
    for r in data:
        instr = str(col(r, "Instructor(s)") or "").strip()
        if instr.lower() == "staff" or not instr:
            staff_count += 1
        else:
            instructors.add(instr.split("/")[0].strip())
    print(f"    Unique named instructors: {len(instructors)}")
    print(f"    Rows with 'Staff' instr:  {staff_count}")

    # -- Previous Semester Days and Times coverage --
    prev_filled = sum(
        1 for rlist in class_rows.values()
        if str(col(rlist[0], "Previous Semester Days and Times") or "").strip()
    )
    print(f"    Prev. semester time set:  {prev_filled}/{unique_sections} sections "
          f"({100 * prev_filled // unique_sections if unique_sections else 0}%)")

    # -- Likely cross-listed groups --
    crosslist_keys: dict[tuple, list] = defaultdict(list)
    for nbr, rlist in class_rows.items():
        r = rlist[0]
        desc = str(col(r, "Description") or "").strip()
        instr = str(col(r, "Instructor(s)") or "").strip().split("/")[0].strip()
        prev = str(col(r, "Previous Semester Days and Times") or "").strip()
        _, cmbnd = _parse_enrollment_str(col(r, "Enrl Cap (Cmbnd Enrl Cap)"))
        crosslist_keys[(desc, instr, prev, cmbnd)].append(nbr)
    crosslisted_groups = {k: v for k, v in crosslist_keys.items() if len(v) > 1}
    crosslisted_sections = sum(len(v) for v in crosslisted_groups.values())
    print(f"    Cross-listed groups:      {len(crosslisted_groups)} "
          f"({crosslisted_sections} sections)")

    # -- Format compatibility --
    print()
    print("    FORMAT COMPATIBILITY")
    missing_cols = _SOC_EDITORS_COLUMNS - set(headers)
    if missing_cols:
        print(f"    ⚠  Missing expected SOC Editors columns: {sorted(missing_cols)}")
    else:
        print("    ✓  All expected SOC Editors columns present")

    _RECOMMENDED_ADDITIONS = [
        "Enrl Tot",
        "Is Core",
        "Is New",
        "Rank",
        "Current Days and Times",
        "Cross-List Parent Class Nbr",
        "Room",
    ]
    present = set(headers)
    missing_recommended = [c for c in _RECOMMENDED_ADDITIONS if c not in present]
    if missing_recommended:
        print(f"    →  Recommended additions for full model coverage: {missing_recommended}")


def _inspect_sis(rows: list[tuple], headers: list[str]) -> None:
    """Print SIS-specific diagnostics (brief)."""
    print(f"    Data rows: {len(rows)}")
    key_cols = ["CLASS_NBR", "INSTR_NAME", "ENRL_CAP", "CLASS_MTG_DAYS", "CW_CLASS_MTG_TIMES"]
    present = [c for c in key_cols if c in headers]
    missing = [c for c in key_cols if c not in headers]
    print(f"    Expected SIS columns present: {present}")
    if missing:
        print(f"    Missing SIS columns: {missing}")


def main() -> None:
    if not INPUT_DIR.exists():
        print(f"Input directory not found: {INPUT_DIR}")
        return

    xlsx_files = list(INPUT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files in {INPUT_DIR}")
        return

    for path in sorted(xlsx_files):
        print(f"\n{'=' * 60}")
        print(f"File: {path.name}")
        print(f"{'=' * 60}")
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  Error loading workbook: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  Sheet: {sheet_name!r}")

            # Read all rows (up to a safe limit for inspection)
            all_rows = list(ws.iter_rows(max_row=600, values_only=True))
            if not all_rows:
                print("    (empty sheet)")
                continue

            # Show raw first 4 rows for manual inspection
            print("  First 4 rows (raw):")
            for i, row in enumerate(all_rows[:4], 1):
                cells = [str(c)[:35] if c is not None else "" for c in row]
                print(f"    Row {i}: {cells}")

            # Find the header row: first row where most cells are non-None strings
            header_row_idx = None
            headers: list[str] = []
            for i, row in enumerate(all_rows):
                non_null = [c for c in row if c is not None]
                if len(non_null) >= 4 and all(isinstance(c, str) for c in non_null[:4]):
                    header_row_idx = i
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    break

            if header_row_idx is None:
                print("    Could not identify header row.")
                continue

            data_rows = all_rows[header_row_idx + 1:]
            data_rows = [r for r in data_rows if any(c is not None for c in r)]

            fmt = _detect_format(headers)
            print(f"\n  Detected format: {fmt.upper().replace('_', ' ')}")
            print(f"  Columns ({len([h for h in headers if h])}):")
            print(f"    {[h for h in headers if h]}")
            print()
            print("  Diagnostics:")

            if fmt == "soc_editors":
                _inspect_soc_editors(data_rows, headers)
            else:
                _inspect_sis(data_rows, headers)

        wb.close()


if __name__ == "__main__":
    main()
