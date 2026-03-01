"""
Convert xlsx input files to SchedulingInput JSON for solver app.py,
OR to a model-compatible JSON for direct ingestion into the platform data model.

Supports two source formats detected automatically from column headers:

  SIS Format  — columns like CLASS_MTG_DAYS, INSTR_NAME, CLASS_NBR
  SOC Editors — columns: Term, Subject, Catalog, Description, Section, Session,
                          Meeting Pattern Number, Previous Semester Days and Times,
                          Meeting Dates, Instructor(s), Enrl Cap (Cmbnd Enrl Cap),
                          Instruction Mode

Usage:
  python convert.py [output_label] [--limit N] [--output-mode solver|model]

Output modes:
  solver  (default) — output/scheduling_input_<label>.json  (solver app.py format)
  model             — output/model_input_<label>.json       (platform data model format)

Demo mode: python convert.py demo --limit 20
"""
import argparse
import json
import re
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

from config import (
    INPUT_DIR,
    OUTPUT_DIR,
    SECTIONS_SOURCE,
    DEFAULT_ROOM_REQUIREMENTS,
    DEFAULT_TAGS,
)

# ============================================================================
# SHARED UTILITIES
# ============================================================================

def _normalize_id(s: str) -> str:
    """Replace spaces and commas with underscores for use as id."""
    if s is None or not isinstance(s, str):
        return ""
    return re.sub(r"[\s,]+", "_", str(s).strip()).strip("_") or "unknown"


def _parse_time_range(s: str) -> tuple[str, str] | None:
    """Parse '6:00 PM-8:30 PM' or '9:00AM-10:15AM' -> ('18:00', '20:30')."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    m = re.match(
        r"(\d{1,2}):(\d{2})\s*(AM|PM)\s*[-–]\s*(\d{1,2}):(\d{2})\s*(AM|PM)",
        s,
        re.IGNORECASE,
    )
    if not m:
        return None
    h1, mn1, ap1, h2, mn2, ap2 = m.groups()
    h1, mn1, h2, mn2 = int(h1), int(mn1), int(h2), int(mn2)
    if ap1.upper() == "PM" and h1 != 12:
        h1 += 12
    if ap1.upper() == "AM" and h1 == 12:
        h1 = 0
    if ap2.upper() == "PM" and h2 != 12:
        h2 += 12
    if ap2.upper() == "AM" and h2 == 12:
        h2 = 0
    return (f"{h1:02d}:{mn1:02d}", f"{h2:02d}:{mn2:02d}")


def _parse_days(s: str) -> list[str]:
    """Parse 'Mon Wed' -> ['Mon', 'Wed']."""
    if not s or not isinstance(s, str):
        return []
    return [p for p in s.strip().split() if len(p) >= 2]


_SOC_DAY_MAP = {
    "M": "Mon", "Tu": "Tue", "T": "Tue",
    "W": "Wed", "Th": "Thu", "R": "Thu",
    "F": "Fri", "Sa": "Sat", "S": "Sat", "Su": "Sun",
}

_DAY_ORDER = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")


def _parse_days_and_times(s: str) -> tuple[list[str], str]:
    """
    Parse SOC-style 'Days and Times' e.g. 'MW 9:00AM - 10:15AM'.
    Returns (days_list, times_str).
    """
    if not s or not isinstance(s, str):
        return [], ""
    s = s.strip()
    if s.upper() == "TBA":
        return [], ""
    m = re.match(r"^([MTWRFSSatuwh]+)\s+(.+)$", s, re.IGNORECASE)
    if not m:
        return [], ""
    day_part, times_part = m.group(1).strip(), m.group(2).strip()
    days: list[str] = []
    i = 0
    while i < len(day_part):
        two = day_part[i: i + 2] if i + 2 <= len(day_part) else ""
        one = day_part[i]
        if two in _SOC_DAY_MAP:
            days.append(_SOC_DAY_MAP[two])
            i += 2
        elif one in _SOC_DAY_MAP:
            days.append(_SOC_DAY_MAP[one])
            i += 1
        else:
            i += 1
    return days, times_part


def _extract_class_nbr_from_section(section_val) -> str | None:
    """Extract class number: '100-LEC (13135)' -> '13135'."""
    if section_val is None:
        return None
    m = re.search(r"\((\d+)\)", str(section_val).strip())
    return m.group(1) if m else None


def _extract_section_code(section_val) -> str:
    """Extract section code: '100-LEC (13135)' -> '100-LEC'."""
    if section_val is None:
        return ""
    m = re.match(r"(\d+-\w+)", str(section_val).strip())
    return m.group(1) if m else str(section_val).strip()


def _extract_section_type(section_code: str) -> str:
    """
    Derive a normalized section_type from the code suffix.
    '100-LEC' -> 'lecture', '400-SEM' -> 'seminar', etc.
    """
    _TYPE_MAP = {
        "LEC": "lecture",
        "SEM": "seminar",
        "LAB": "lab",
        "IND": "individualized_study",
        "DSR": "dissertation",
        "PRA": "practicum",
        "RSC": "research",
        "WRK": "workshop",
    }
    m = re.search(r"-(\w+)$", section_code)
    if m:
        return _TYPE_MAP.get(m.group(1).upper(), m.group(1).lower())
    return "lecture"


def _parse_enrollment_str(val) -> tuple[int, int]:
    """
    Parse 'section_cap (combined_cap)' e.g. '21 (45)' -> (21, 45).
    Returns (section_cap, combined_cap). combined_cap defaults to section_cap if absent.
    """
    if val is None:
        return 0, 0
    s = str(val).strip()
    m = re.match(r"^(\d+)\s*\((\d+)\)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = re.match(r"^(\d+)", s)
    if m2:
        v = int(m2.group(1))
        return v, v
    return 0, 0


def _parse_enrollment_val(val) -> int:
    """Parse single enrollment value; handles '35 (38)' by taking first number."""
    return _parse_enrollment_str(val)[0]


def _parse_instruction_mode(val) -> tuple[str, str]:
    """
    Parse 'P (In Person, Traditional)' -> ('P', 'In Person, Traditional').
    """
    if not val:
        return "", ""
    s = str(val).strip()
    m = re.match(r"^(\w+)\s*\((.+)\)\s*$", s)
    if m:
        return m.group(1), m.group(2)
    return s, s


def _build_timeslot_id(day: str, start: str, end: str) -> str:
    return f"{day}_{start}_{end}".replace(":", "")


def _sort_slot_ids_by_day_order(slot_ids: list[str]) -> list[str]:
    def day_rank(slot_id: str) -> int:
        for i, day in enumerate(_DAY_ORDER):
            if slot_id.startswith(day + "_"):
                return i
        return 999
    return sorted(slot_ids, key=day_rank)


def _parse_room_and_capacity(s: str) -> tuple[str, int | None]:
    if not s or not isinstance(s, str):
        return "TBA", None
    s = s.strip()
    if not s or s.upper() == "TO BE ANNOUNCED":
        return "TBA", None
    m = re.match(r"^(.+?)\s*\((\d+)\)\s*$", s)
    if m:
        return m.group(1).strip(), int(m.group(2))
    return s, None


def _read_sheet_with_headers(
    path: Path,
    sheet_name: str,
    header_row: int,
    data_start: int,
    max_column: int | None = None,
) -> tuple[list[str], list[dict]]:
    """Return (column_names, row_dicts)."""
    use_read_only = max_column is None
    wb = openpyxl.load_workbook(path, read_only=use_read_only, data_only=True)
    ws = wb[sheet_name]
    kwargs: dict = {"min_row": header_row, "max_row": header_row, "values_only": True}
    if max_column is not None:
        kwargs["max_col"] = max_column
    header_cells = next(ws.iter_rows(**kwargs))
    headers = [
        (i + 1, str(v).strip().replace("\n", " ").strip() if v is not None else "")
        for i, v in enumerate(header_cells)
    ]
    headers = [(col, name or None) for col, name in headers]
    while headers and headers[-1][1] is None:
        headers.pop()
    row_kwargs: dict = {"min_row": data_start, "values_only": True}
    if max_column is not None:
        row_kwargs["max_col"] = max_column
    rows = []
    for row in ws.iter_rows(**row_kwargs):
        row_dict = {name: val for (_, name), val in zip(headers, row) if name}
        rows.append(row_dict)
    wb.close()
    return [name for _, name in headers if name], rows


# ============================================================================
# FORMAT DETECTION
# ============================================================================

# Canonical column names for the SOC Editors export format
_SOC_EDITORS_COLUMNS = {
    "Term", "Subject", "Catalog", "Description", "Section", "Session",
    "Meeting Pattern Number", "Previous Semester Days and Times",
    "Meeting Dates", "Instructor(s)", "Enrl Cap (Cmbnd Enrl Cap)", "Instruction Mode",
}


def detect_format(path: Path, sheet_name: str, header_row: int) -> str:
    """
    Return 'soc_editors' or 'sis' based on column headers in the file.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    wb.close()
    found = {str(v).strip() for v in header_cells if v is not None}
    overlap = found & _SOC_EDITORS_COLUMNS
    return "soc_editors" if len(overlap) >= 6 else "sis"


# ============================================================================
# SOC EDITORS FORMAT — LOADER
# ============================================================================

def load_sections_from_soc_editors(
    max_rows: int | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], dict]:
    """
    Load the SOC Editors xlsx format (testing_updated.xlsx style) and return:
      (sections, instructors, rooms, timeslots, meeting_patterns, gaps_report)

    Row deduplication: multiple rows with the same class number represent
    separate meeting occurrences (Meeting Pattern Number 1…N) for ONE section.
    They are collapsed into a single section record; the count of occurrences
    is stored as meeting_sessions_count.

    Returns model-compatible dicts aligned with model.py field names.
    """
    path = INPUT_DIR / SECTIONS_SOURCE["file"]
    if not path.exists():
        return [], [], [], [], [], {}

    _, rows = _read_sheet_with_headers(
        path,
        SECTIONS_SOURCE["sheet"],
        SECTIONS_SOURCE["header_row"],
        SECTIONS_SOURCE["data_start_row"],
        max_column=SECTIONS_SOURCE.get("max_column"),
    )
    if max_rows is not None:
        rows = rows[:max_rows]

    # -- Deduplicate: group rows by class number ----------------------------
    # Each unique class number = one section. Multiple rows = meeting occurrences.
    class_rows: dict[str, list[dict]] = defaultdict(list)
    rows_missing_class_nbr = 0

    for row in rows:
        class_nbr = _extract_class_nbr_from_section(row.get("Section"))
        if class_nbr is None:
            rows_missing_class_nbr += 1
            continue
        class_rows[class_nbr].append(row)

    # -- Entity accumulators -----------------------------------------------
    timeslot_id_to_timeslot: dict[str, dict] = {}
    pattern_key_to_pattern: dict[str, dict] = {}
    instr_id_to_instructor: dict[str, dict] = {}

    sections: list[dict] = []

    for class_nbr, occurrences in class_rows.items():
        # Take the first occurrence as the canonical row; all occurrences are
        # identical except for Meeting Pattern Number.
        row = occurrences[0]
        meeting_sessions_count = max(
            int(r.get("Meeting Pattern Number") or 1) for r in occurrences
        )

        # -- Course fields --------------------------------------------------
        subject = str(row.get("Subject") or "").strip()
        catalog = str(row.get("Catalog") or "").strip()
        description = str(row.get("Description") or "").strip()
        course_id = _normalize_id(f"{subject}{catalog}") or class_nbr

        # -- Section fields -------------------------------------------------
        section_raw = row.get("Section") or ""
        section_code = _extract_section_code(section_raw)
        section_type = _extract_section_type(section_code)

        # -- Enrollment -----------------------------------------------------
        section_cap, combined_cap = _parse_enrollment_str(
            row.get("Enrl Cap (Cmbnd Enrl Cap)")
        )

        # -- Instruction mode -----------------------------------------------
        mode_code, _mode_label = _parse_instruction_mode(row.get("Instruction Mode"))
        is_virtual = mode_code in ("DT",)          # DT = Distance Total (fully online)
        is_hybrid = mode_code in ("HY",)

        # -- Session / Meeting Dates ----------------------------------------
        session = str(row.get("Session") or "").strip()
        meeting_dates = str(row.get("Meeting Dates") or "").strip()

        # -- Instructor -----------------------------------------------------
        instr_raw = str(row.get("Instructor(s)") or "").strip()
        # Take first instructor if multiple separated by /
        instr_name = instr_raw.split("/")[0].strip()
        is_staff = instr_name.lower() == "staff" or not instr_name
        instr_id = _normalize_id(instr_name) if not is_staff else "staff_tbd"
        if instr_id not in instr_id_to_instructor:
            instr_id_to_instructor[instr_id] = {
                "id": instr_id,
                "name": instr_name if not is_staff else "Staff TBD",
                # GAP: rank_type not present in source — must be set manually
                "rank_type": None,
                "unavailable_times": [],
                "preferences": {
                    "preferred_times": [],
                    "preferred_days": [],
                    "preferred_patterns": [],
                    "unavailable_times": [],
                    "max_teaching_days": None,
                },
            }

        # -- Previous Semester Days and Times -> last_year_time + pattern ----
        prev_days_times = str(row.get("Previous Semester Days and Times") or "").strip()
        pattern_id: str | None = None
        if prev_days_times:
            days, times_str = _parse_days_and_times(prev_days_times)
            time_range = _parse_time_range(times_str)
            if days and time_range:
                start_time, end_time = time_range
                slot_ids = []
                for day in days:
                    slot_id = _build_timeslot_id(day, start_time, end_time)
                    slot_ids.append(slot_id)
                    if slot_id not in timeslot_id_to_timeslot:
                        timeslot_id_to_timeslot[slot_id] = {
                            "id": slot_id,
                            "day": day,
                            "start_time": start_time,
                            "end_time": end_time,
                            # GAP: slot_type (standard/evening/lab) not derivable
                            #      without a defined time boundary policy.
                            "slot_type": None,
                        }
                ordered_slots = _sort_slot_ids_by_day_order(slot_ids)
                pattern_key = "|".join(ordered_slots) + "|"
                if pattern_key not in pattern_key_to_pattern:
                    pattern_key_to_pattern[pattern_key] = {
                        "id": _normalize_id(pattern_key) or "default",
                        "slots_required": len(slot_ids),
                        "allowed_days": days,
                        "compatible_timeslot_sets": [slot_ids],
                    }
                pattern_id = pattern_key_to_pattern[pattern_key]["id"]

        # -- Build section dict aligned with model.py Section ---------------
        sections.append({
            # Core identifiers
            "id": class_nbr,
            "course_id": course_id,
            "section_code": section_code,
            "section_type": section_type,

            # Instructor
            "instructor_id": instr_id,

            # Enrollment
            "enrollment_cap": section_cap,
            "combined_enrollment_cap": combined_cap,   # combined across cross-listed sections
            # GAP: expected_enrollment (actual enrolled count) not in source —
            #      using enrollment_cap as proxy until actuals are available.
            "expected_enrollment": section_cap,

            # Assignment — not yet assigned; these will be set by the solver
            "room_id": None,           # GAP: no room data in this file
            "timeslot_id": None,       # GAP: current assignment not yet made

            # Cross-listing — detected separately, resolved in post-processing
            "crosslisting_id": None,   # GAP: set after crosslist_groups are built
            "crosslist_group_id": None,

            # Historical reference
            "last_year_time": prev_days_times or None,
            "last_year_room": None,    # GAP: not present in source file

            # Scheduling constraints
            "allowed_meeting_patterns": [pattern_id] if pattern_id else [],
            "room_requirements": DEFAULT_ROOM_REQUIREMENTS.copy(),
            "tags": DEFAULT_TAGS.copy(),

            # Supplemental metadata (not in core model, useful for downstream logic)
            "_meta": {
                "term": str(row.get("Term") or "").strip(),
                "subject": subject,
                "catalog": catalog,
                "description": description,
                "session": session,
                "meeting_dates": meeting_dates,
                "meeting_sessions_count": meeting_sessions_count,
                "instruction_mode_code": mode_code,
                "is_virtual": is_virtual,
                "is_hybrid": is_hybrid,
                "is_crosslisted": False,   # updated in post-processing
            },
        })

    # -------------------------------------------------------------------------
    # POST-PROCESSING: detect cross-listed groups
    # -------------------------------------------------------------------------
    # Signal: sections sharing the same (description, instructor, last_year_time,
    #         combined_enrollment_cap) are almost certainly cross-listed.
    #         Groups of 1 are standalone.
    crosslist_key: dict[tuple, list[str]] = defaultdict(list)
    for s in sections:
        key = (
            s["_meta"]["description"],
            s["instructor_id"],
            s.get("last_year_time") or "",
            s["combined_enrollment_cap"],
        )
        crosslist_key[key].append(s["id"])

    crosslist_groups: list[dict] = []
    section_by_id = {s["id"]: s for s in sections}

    for key, ids in crosslist_key.items():
        if len(ids) > 1:
            group_id = _normalize_id(f"xlist_{'_'.join(sorted(ids))}")
            crosslist_groups.append({
                "id": group_id,
                "member_section_ids": sorted(ids),
                "require_same_room": True,
            })
            for sid in ids:
                section_by_id[sid]["crosslist_group_id"] = group_id
                section_by_id[sid]["_meta"]["is_crosslisted"] = True

    # -------------------------------------------------------------------------
    # POST-PROCESSING: expand allowed_meeting_patterns by slot count
    # -------------------------------------------------------------------------
    pattern_id_to_slots = {p["id"]: p["slots_required"] for p in pattern_key_to_pattern.values()}
    all_pattern_ids = list(pattern_id_to_slots.keys())
    for s in sections:
        prev_pattern = (s["allowed_meeting_patterns"] or [None])[0]
        if prev_pattern and prev_pattern in pattern_id_to_slots:
            n = pattern_id_to_slots[prev_pattern]
            s["allowed_meeting_patterns"] = [
                pid for pid in all_pattern_ids if pattern_id_to_slots[pid] == n
            ]
        else:
            s["allowed_meeting_patterns"] = all_pattern_ids

    # -------------------------------------------------------------------------
    # Gaps report
    # -------------------------------------------------------------------------
    gaps_report = _build_gaps_report(sections, rows_missing_class_nbr)

    timeslots = list(timeslot_id_to_timeslot.values())
    meeting_patterns = list(pattern_key_to_pattern.values())
    instructors = list(instr_id_to_instructor.values())

    # Rooms: empty — source file has no room data
    rooms: list[dict] = []

    return sections, instructors, rooms, timeslots, meeting_patterns, gaps_report


def _build_gaps_report(sections: list[dict], rows_missing_class_nbr: int) -> dict:
    """
    Document every field that could not be populated from the source file
    and suggest remediation strategies.
    """
    staff_sections = [s for s in sections if s["instructor_id"] == "staff_tbd"]
    crosslisted_count = sum(1 for s in sections if s["_meta"]["is_crosslisted"])

    return {
        "summary": {
            "total_sections": len(sections),
            "rows_dropped_missing_class_nbr": rows_missing_class_nbr,
            "sections_with_no_previous_pattern": sum(
                1 for s in sections if not s.get("last_year_time")
            ),
            "cross_listed_sections": crosslisted_count,
            "staff_tbd_instructors": len(staff_sections),
        },
        "model_fields_not_populated": {
            # ---- Section fields ----
            "Section.room_id": {
                "reason": "Source file contains no room assignment column.",
                "fix_option_1": "Add 'Room' column to the spreadsheet with room ID or name.",
                "fix_option_2": "Run solver to auto-assign rooms; update Section.room_id from results.",
            },
            "Section.timeslot_id": {
                "reason": "File captures *previous* semester times, not current assignment.",
                "fix_option_1": "Run solver; Section.timeslot_id is set from solver output.",
                "fix_option_2": "If current semester times are already confirmed, add 'Current Days and Times' column.",
            },
            "Section.crosslisting_id": {
                "reason": "Self-referential FK requires a canonical 'parent' section to be designated within each group.",
                "fix_option_1": "After import, pick one member_section_id per CrossListGroup as the parent and backfill crosslisting_id.",
                "fix_option_2": "Add 'Cross-List Parent Class Nbr' column to spreadsheet.",
            },
            "Section.last_year_room": {
                "reason": "Previous semester room not present in this export.",
                "fix_option_1": "Join against the prior semester SOC Results export (CW_SR_SOC_SUMMARY_RESULTS) on class number.",
            },
            "Section.expected_enrollment": {
                "reason": "Source only contains enrollment CAP, not actual enrolled count.",
                "note": "Currently defaulting to section_cap as a proxy.",
                "fix_option_1": "Add 'Enrl Tot' column (actual enrollment) to the spreadsheet export.",
            },
            # ---- Instructor fields ----
            "Instructor.rank_type": {
                "reason": "Rank (Faculty / Adjunct / Staff) not present in this export.",
                "fix_option_1": "Add 'Rank' column to spreadsheet.",
                "fix_option_2": "Maintain a separate instructor roster CSV and join on name.",
            },
            "InstructorPreferences.*": {
                "reason": "Instructor preferences (preferred days, max teaching days, etc.) are not in any spreadsheet column.",
                "fix_option_1": "Populate InstructorPreferences manually via admin UI after import.",
                "fix_option_2": "Add a separate instructor preferences sheet or CSV.",
            },
            # ---- Room / Timeslot fields ----
            "Room.*": {
                "reason": "No room data exists in the source file.",
                "fix_option_1": "Maintain a separate rooms master list and import independently.",
            },
            "Timeslot.slot_type": {
                "reason": "Cannot classify standard/evening/lab without a defined time boundary policy.",
                "fix_option_1": "Define time boundary rules in config (e.g., evening = start >= 17:00) and auto-classify.",
            },
            # ---- Course fields ----
            "Course.is_core": {
                "reason": "Core course flag not present in this export.",
                "fix_option_1": "Add 'Is Core' (Y/N) column to spreadsheet.",
                "fix_option_2": "Maintain a separate core courses list by Subject+Catalog.",
            },
            "Course.is_new": {
                "reason": "New course flag not present in this export.",
                "fix_option_1": "Add 'Is New' (Y/N) column to spreadsheet.",
            },
            # ---- Other model tables ----
            "Major.*": {
                "reason": "No major/program data in this file.",
                "fix_option_1": "Import separately from curriculum system.",
            },
            "DepartmentPreferences.*": {
                "reason": "Department-level collision and virtual preferences not in this file.",
                "fix_option_1": "Populate manually via admin UI per department.",
            },
            "RoomPreferences.*": {
                "reason": "Room-specific needs (projector, lab, etc.) not derivable from this file.",
                "fix_option_1": "Populate from rooms master list or admin UI.",
            },
            "SectionPreferences.cannot_collide_with": {
                "reason": "No non-overlap constraints expressed in this file.",
                "fix_option_1": "Add 'Cannot Overlap With (Class Nbrs)' column to spreadsheet.",
            },
        },
        "spreadsheet_columns_not_mapped_to_model": {
            "Term": {
                "value_example": "2268 - Fall 2026",
                "note": "Useful for multi-term filtering but no Term table in model.",
                "fix_option_1": "Add a Term table to model.py; FK from Section.",
                "fix_option_2": "Store as a tag on each Section.",
            },
            "Session": {
                "value_example": "Dyn Dt, Regular, Reg 2 Half, Virt Reg, Virt Sess1, Virt Sess2",
                "note": "Indicates sub-semester session windows and delivery format.",
                "fix_option_1": "Add Session field to Section model. Virtual sessions can also set Section.allow_virtual.",
                "fix_option_2": "Map 'Reg 2 Half', 'Virt Sess1/2' to BlockedTime constraints restricting placement windows.",
            },
            "Meeting Dates": {
                "value_example": "8/24/2026 - 12/7/2026",
                "note": "Date range for this meeting occurrence. Differs across half-semester sessions.",
                "fix_option_1": "Add start_date / end_date columns to Section or create a Session table.",
            },
            "Meeting Pattern Number": {
                "value_example": "1 … 11",
                "note": "Count of individual meeting occurrences (e.g., 6 Saturday classes). "
                        "Collapsed to 'meeting_sessions_count' in _meta.",
                "fix_option_1": "Add meeting_sessions_count (Integer) to Section model for scheduling weight.",
            },
        },
        "cross_listing_note": (
            "Cross-listing was inferred from matching (Description, Instructor, "
            "Previous Days/Times, Combined Enrollment Cap). This heuristic is "
            f"reliable but not authoritative — {crosslisted_count} sections were "
            "grouped. Review CrossListGroup records before relying on them."
        ),
        "staff_tbd_note": (
            f"{len(staff_sections)} sections list 'Staff' as instructor. "
            "These are assigned instructor_id='staff_tbd' and must be updated "
            "with a real instructor before scheduling."
        ),
    }


# ============================================================================
# SIS FORMAT — original loader (unchanged from prior version)
# ============================================================================

def load_sections_from_sis(
    max_rows: int | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict]]:
    """
    Load SIS Schedule xlsx and return (sections, instructors, rooms, timeslots, meeting_patterns).
    Produces solver-compatible JSON (original behavior).
    """
    path = INPUT_DIR / SECTIONS_SOURCE["file"]
    if not path.exists():
        return [], [], [], [], []

    max_col = SECTIONS_SOURCE.get("max_column")
    _, rows = _read_sheet_with_headers(
        path,
        SECTIONS_SOURCE["sheet"],
        SECTIONS_SOURCE["header_row"],
        SECTIONS_SOURCE["data_start_row"],
        max_column=max_col,
    )
    if max_rows is not None:
        rows = rows[:max_rows]

    timeslot_id_to_timeslot: dict[str, dict] = {}
    pattern_key_to_pattern: dict[str, dict] = {}
    room_name_to_room: dict[str, dict] = {}
    instr_name_to_instructor: dict[str, dict] = {}
    sections: list[dict] = []

    for row in rows:
        days_str = row.get("CLASS_MTG_DAYS") or row.get("days") or ""
        times_str = row.get("CW_CLASS_MTG_TIMES") or row.get("times") or ""
        days_and_times = row.get("Days and Times") or ""
        if days_and_times:
            days, times_str = _parse_days_and_times(days_and_times)
        else:
            days = _parse_days(days_str)
        time_range = _parse_time_range(times_str)

        pattern_id: str | None = None
        if days and time_range:
            start_time, end_time = time_range
            pattern_slot_ids: list[str] = []
            for day in days:
                slot_id = _build_timeslot_id(day, start_time, end_time)
                pattern_slot_ids.append(slot_id)
                if slot_id not in timeslot_id_to_timeslot:
                    timeslot_id_to_timeslot[slot_id] = {
                        "id": slot_id, "day": day,
                        "start_time": start_time, "end_time": end_time,
                    }
            ordered_slots = _sort_slot_ids_by_day_order(pattern_slot_ids)
            pattern_key = "|".join(ordered_slots) + "|"
            if pattern_key not in pattern_key_to_pattern:
                pattern_key_to_pattern[pattern_key] = {
                    "id": _normalize_id(pattern_key) or "default",
                    "slots_required": len(pattern_slot_ids),
                    "allowed_days": list(days),
                    "compatible_timeslot_sets": [pattern_slot_ids],
                }
            pattern_id = pattern_key_to_pattern[pattern_key]["id"]

        room_raw = row.get("CW_MEETING_ROOM") or row.get("Room (Capacity)") or row.get("room") or "TBA"
        room_str, room_cap_from_col = _parse_room_and_capacity(
            room_raw if isinstance(room_raw, str) else str(room_raw or "")
        )
        if not room_str or room_str.upper() == "TO BE ANNOUNCED":
            room_str = "TBA"
        if room_str not in room_name_to_room:
            cap_val = room_cap_from_col or row.get("ENRL_CAP") or row.get("enrollment_cap")
            try:
                cap = int(float(cap_val)) if cap_val is not None else 50
            except (TypeError, ValueError):
                cap = 50
            room_name_to_room[room_str] = {
                "id": _normalize_id(room_str) or "TBA",
                "building": room_str,
                "capacity": max(cap, 10),
                "features": [],
            }

        instr_raw = row.get("INSTR_NAME") or row.get("Instructor(s)") or row.get("Instructor") or ""
        instr_name = str(instr_raw or "").strip()
        if instr_name and instr_name.upper() != "TO BE ANNOUNCED":
            instr_name = instr_name.split("/")[0].split(";")[0].strip()
        if instr_name and instr_name not in instr_name_to_instructor:
            instr_name_to_instructor[instr_name] = {
                "id": _normalize_id(instr_name) or "unknown",
                "rank_type": "Faculty",
                "unavailable_times": [],
                "preferences": {
                    "preferred_days": [], "preferred_patterns": [], "max_teaching_days": None,
                },
            }

        class_nbr = row.get("CLASS_NBR") or row.get("class_nbr")
        if class_nbr is None:
            class_nbr = _extract_class_nbr_from_section(row.get("Section"))
        if class_nbr is None:
            continue
        section_id = str(int(float(class_nbr))) if isinstance(class_nbr, (int, float)) else str(class_nbr).strip()
        subject = (row.get("SUBJECT") or row.get("Subject") or "").strip()
        catalog_nbr = (row.get("CATALOG_NBR") or row.get("catalog_nbr") or "").strip()
        if not catalog_nbr:
            desc = (row.get("Description") or "").strip()
            if desc:
                first_part = desc.split(" - ", 1)[0].split()
                catalog_nbr = first_part[-1] if first_part else ""
        course_id = _normalize_id(f"{subject}{catalog_nbr}") or section_id
        section_code = (row.get("CLASS_SECTION") or row.get("Section") or "").strip() or section_id
        enrl_cap = _parse_enrollment_val(row.get("ENRL_CAP") or row.get("enrollment_cap"))
        enrl_tot = _parse_enrollment_val(row.get("ENRL_TOT") or row.get("enrollment_total"))
        instr_id = _normalize_id(instr_name) if instr_name else "unknown"
        room_id = room_name_to_room[room_str]["id"]

        sections.append({
            "id": section_id, "course_id": course_id, "section_code": section_code,
            "instructor_id": instr_id, "expected_enrollment": enrl_tot,
            "enrollment_cap": enrl_cap,
            "allowed_meeting_patterns": [pattern_id] if pattern_id is not None else [],
            "previous_meeting_pattern": pattern_id,
            "room_requirements": DEFAULT_ROOM_REQUIREMENTS.copy(),
            "crosslist_group_id": None, "tags": DEFAULT_TAGS.copy(),
        })

    pattern_id_to_slots = {p["id"]: p["slots_required"] for p in pattern_key_to_pattern.values()}
    all_pattern_ids = list(pattern_id_to_slots.keys())
    for s in sections:
        prev_id = s.get("previous_meeting_pattern")
        if prev_id and prev_id in pattern_id_to_slots:
            n = pattern_id_to_slots[prev_id]
            s["allowed_meeting_patterns"] = [
                pid for pid in all_pattern_ids if pattern_id_to_slots[pid] == n
            ]
        else:
            s["allowed_meeting_patterns"] = all_pattern_ids

    return (
        sections, list(instr_name_to_instructor.values()),
        list(room_name_to_room.values()),
        list(timeslot_id_to_timeslot.values()),
        list(pattern_key_to_pattern.values()),
    )


# ============================================================================
# OUTPUT BUILDERS
# ============================================================================

def build_scheduling_input(max_rows: int | None = None) -> dict:
    """Solver-compatible SchedulingInput JSON (original SIS format path)."""
    sections, instructors, rooms, timeslots, meeting_patterns = load_sections_from_sis(
        max_rows=max_rows
    )
    return {
        "sections": sections, "instructors": instructors, "rooms": rooms,
        "timeslots": timeslots, "meeting_patterns": meeting_patterns,
        "crosslist_groups": [], "no_overlap_groups": [], "blocked_times": [],
        "locked_assignments": [], "soft_locks": [],
    }


def build_model_input(max_rows: int | None = None) -> dict:
    """
    Model-compatible JSON from SOC Editors format.
    Fields align with model.py entity names and types.
    Includes a '_gaps' section documenting what could not be populated.
    """
    sections, instructors, rooms, timeslots, meeting_patterns, gaps = (
        load_sections_from_soc_editors(max_rows=max_rows)
    )

    # Derive unique courses from section metadata
    courses: dict[str, dict] = {}
    for s in sections:
        cid = s["course_id"]
        if cid not in courses:
            meta = s.get("_meta", {})
            courses[cid] = {
                "id": cid,
                "title": meta.get("description", ""),
                "department": meta.get("subject", ""),
                # GAP: is_core and is_new not in source
                "is_core": None,
                "is_new": None,
            }

    # Strip _meta from sections before final output (it's for internal use only)
    clean_sections = []
    for s in sections:
        cs = {k: v for k, v in s.items() if k != "_meta"}
        clean_sections.append(cs)

    # Collect crosslist_groups from section data
    seen_groups: set[str] = set()
    crosslist_groups: list[dict] = []
    for s in sections:
        gid = s.get("crosslist_group_id")
        if gid and gid not in seen_groups:
            seen_groups.add(gid)
            member_ids = [
                other["id"] for other in sections
                if other.get("crosslist_group_id") == gid
            ]
            crosslist_groups.append({
                "id": gid,
                "member_section_ids": sorted(member_ids),
                "require_same_room": True,
            })

    return {
        "courses": list(courses.values()),
        "instructors": instructors,
        "sections": clean_sections,
        "rooms": rooms,                    # empty — not in source
        "timeslots": timeslots,            # derived from previous semester
        "meeting_patterns": meeting_patterns,
        "crosslist_groups": crosslist_groups,
        "no_overlap_groups": [],           # GAP: not in source
        "blocked_times": [],               # GAP: not in source
        "locked_assignments": [],          # GAP: not in source
        "soft_locks": [],                  # GAP: not in source
        "_gaps": gaps,
    }


# ============================================================================
# VALIDATION
# ============================================================================

def validate_with_app(data: dict) -> None:
    """Validate solver JSON against app.SchedulingInput Pydantic model (if available)."""
    solver_dir = Path(__file__).resolve().parent.parent
    if str(solver_dir) not in sys.path:
        sys.path.insert(0, str(solver_dir))
    try:
        from app import SchedulingInput
        SchedulingInput(**data)
    except ImportError:
        pass
    except Exception as e:
        raise ValueError(f"SchedulingInput validation failed: {e}") from e


# ============================================================================
# CLI
# ============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert xlsx schedule input to JSON.",
        epilog="Example (demo): python convert.py demo --limit 20 --output-mode model",
    )
    parser.add_argument("label", nargs="?", default="fall2026",
                        help="Output file label")
    parser.add_argument("--limit", type=int, default=None, metavar="N",
                        help="Process only first N data rows")
    parser.add_argument(
        "--output-mode", choices=["solver", "model"], default=None,
        help="Output JSON format. Defaults: 'model' for SOC Editors format, "
             "'solver' for SIS format.",
    )
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    path = INPUT_DIR / SECTIONS_SOURCE["file"]
    fmt = detect_format(path, SECTIONS_SOURCE["sheet"], SECTIONS_SOURCE["header_row"])
    output_mode = args.output_mode or ("model" if fmt == "soc_editors" else "solver")

    print(f"Detected format: {fmt}  →  output mode: {output_mode}")

    if output_mode == "model":
        out_path = OUTPUT_DIR / f"model_input_{args.label}.json"
        data = build_model_input(max_rows=args.limit)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        gaps = data["_gaps"]["summary"]
        limit_msg = f" (limited to {args.limit} rows)" if args.limit else ""
        print(
            f"Wrote {len(data['sections'])} sections, "
            f"{len(data['instructors'])} instructors, "
            f"{len(data['courses'])} courses, "
            f"{len(data['crosslist_groups'])} cross-list groups to {out_path}{limit_msg}"
        )
        print(f"Gaps: {gaps['sections_with_no_previous_pattern']} sections missing previous pattern, "
              f"{gaps['staff_tbd_instructors']} staff TBD, "
              f"{gaps['cross_listed_sections']} cross-listed.")
        print(f"Review _gaps in output JSON for full remediation guide.")
    else:
        out_path = OUTPUT_DIR / f"scheduling_input_{args.label}.json"
        data = build_scheduling_input(max_rows=args.limit)
        validate_with_app(data)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        limit_msg = f" (limited to {args.limit} rows)" if args.limit else ""
        print(
            f"Wrote {len(data['sections'])} sections, {len(data['instructors'])} instructors, "
            f"{len(data['rooms'])} rooms, {len(data['timeslots'])} timeslots, "
            f"{len(data['meeting_patterns'])} meeting patterns to {out_path}{limit_msg}"
        )


if __name__ == "__main__":
    main()

def _classify_slot_type(start_time: str, section_type: str | None = None) -> str:
    if section_type and section_type.upper() in LAB_SECTION_TYPES:
        return "lab"
    for slot_type, bounds in TIMESLOT_CLASSIFICATION.items():
        if bounds["start_min"] <= start_time <= bounds["start_max"]:
            return slot_type
    return "standard"